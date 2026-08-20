from __future__ import annotations

import html as html_entities
import io
import re
import zipfile
from dataclasses import dataclass, field
from html.parser import HTMLParser
from urllib.parse import urlparse

ZIP_MAGIC = b"PK\x03\x04"
PDF_MAGIC = b"%PDF"
OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"
RAR_MAGIC = b"Rar!\x1a\x07"


@dataclass(frozen=True)
class ParsedDocument:
    text: str
    rows: list[list[str]] = field(default_factory=list)
    parser_status: str = "parsed"
    warnings: list[str] = field(default_factory=list)
    sheets: list[tuple[str, list[list[str]]]] = field(default_factory=list)


class _TextParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.parts.append(data.strip())


def parse_document(body: bytes, content_type: str, url: str = "", charset: str | None = None) -> ParsedDocument:
    normalized_type = (content_type or "").lower()
    suffix = urlparse(url).path.lower()
    if normalized_type in {"text/html", "application/xhtml+xml"} or suffix.endswith((".html", ".htm")):
        parser = _TextParser()
        parser.feed(_decode_text(body, charset))
        return ParsedDocument("\n".join(parser.parts))
    if normalized_type == "application/pdf" or suffix.endswith(".pdf") or body.startswith(PDF_MAGIC):
        return _parse_pdf(body)
    if "spreadsheet" in normalized_type or "excel" in normalized_type or suffix.endswith((".xlsx", ".xlsm", ".xls")):
        return _parse_xlsx(body)
    if "wordprocessingml" in normalized_type or suffix.endswith(".docx"):
        return _parse_docx(body)
    if normalized_type == "application/msword" or suffix.endswith(".doc"):
        if body.startswith(ZIP_MAGIC):
            return _parse_docx(body, note="服务器标记为 msword，实际为 OOXML 压缩包")
        return _parse_ole2(body)
    # 服务端未给出有效类型时，按文件魔数嗅探（ZIP 容器多为 xlsx/docx）。
    if body.startswith(ZIP_MAGIC):
        return _parse_zip_container(body, content_type)
    if body.startswith(OLE2_MAGIC):
        return _parse_ole2(body)
    if body.startswith(RAR_MAGIC):
        return ParsedDocument("", parser_status="unsupported_format", warnings=["RAR 压缩包无法在线解包，请下载附件人工查看"])
    if _looks_binary(body):
        return ParsedDocument("", parser_status="binary_skipped", warnings=[f"二进制附件未解析为文本: {content_type or 'unknown'}"])
    return ParsedDocument(body.decode("utf-8", errors="replace"), parser_status="unknown_type", warnings=[f"unsupported content type: {content_type or 'unknown'}"])


def _parse_pdf(body: bytes) -> ParsedDocument:
    try:
        import fitz  # type: ignore
    except ImportError:
        return ParsedDocument("", parser_status="needs_dependency", warnings=["PDF parsing requires PyMuPDF (fitz)"])
    try:
        document = fitz.open(stream=body, filetype="pdf")
        text = "\n".join(page.get_text() for page in document).strip()
    except Exception as exc:
        return ParsedDocument("", parser_status="error", warnings=[f"PDF 解析失败：{exc}"])
    return ParsedDocument(text, parser_status="parsed" if text else "empty", warnings=[] if text else ["PDF has no extractable text layer"])


def _parse_xlsx(body: bytes, note: str = "") -> ParsedDocument:
    warnings = [note] if note else []
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return ParsedDocument("", parser_status="needs_dependency", warnings=warnings + ["XLSX parsing requires openpyxl"])
    try:
        workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    except Exception as exc:
        return _parse_xls(body, f"openpyxl 无法解析，尝试 xlrd（{exc}）")
    sheets: list[tuple[str, list[list[str]]]] = []
    rows: list[list[str]] = []
    for sheet in workbook.worksheets:
        sheet_rows: list[list[str]] = []
        for values in sheet.iter_rows(values_only=True):
            row = [str(value).strip() if value is not None else "" for value in values]
            if any(row):
                sheet_rows.append(row)
        if sheet_rows:
            sheets.append((sheet.title, sheet_rows))
            rows.extend(sheet_rows)
    text = "\n".join("\t".join(row) for row in rows)
    return ParsedDocument(text, rows=rows, sheets=sheets, warnings=warnings)


def _parse_docx(body: bytes, note: str = "") -> ParsedDocument:
    warnings = [note] if note else []
    try:
        with zipfile.ZipFile(io.BytesIO(body)) as archive:
            if "word/document.xml" not in archive.namelist():
                return ParsedDocument("", parser_status="unsupported_format", warnings=warnings + ["ZIP 包缺少 word/document.xml，不是 DOCX"])
            xml = archive.read("word/document.xml").decode("utf-8", "ignore")
    except Exception as exc:
        return ParsedDocument("", parser_status="error", warnings=warnings + [f"DOCX 解析失败：{exc}"])
    xml = re.sub(r"<w:tab[^>]*/>", "\t", xml)
    xml = re.sub(r"</w:p>", "\n", xml)
    text = re.sub(r"<[^>]+>", "", xml)
    text = html_entities.unescape(text)
    text = "\n".join(line.strip() for line in text.splitlines() if line.strip())
    return ParsedDocument(text, parser_status="parsed" if text else "empty", warnings=warnings)


def _parse_zip_container(body: bytes, content_type: str) -> ParsedDocument:
    spreadsheet = _parse_xlsx(body, note=f"按魔数识别为 ZIP 容器（声明类型: {content_type or 'unknown'}）")
    if spreadsheet.parser_status in {"parsed", "empty"} and (spreadsheet.rows or spreadsheet.text):
        return spreadsheet
    word = _parse_docx(body)
    if word.parser_status in {"parsed", "empty"} and word.text:
        return ParsedDocument(word.text, parser_status=word.parser_status, warnings=[f"按魔数识别为 DOCX（声明类型: {content_type or 'unknown'}）"])
    nested = _parse_zip_entries(body)
    if nested:
        return nested
    return ParsedDocument("", parser_status="unsupported_format", warnings=[f"ZIP 容器无法解析为 xlsx/docx（声明类型: {content_type or 'unknown'}）"])


def _parse_zip_entries(body: bytes) -> ParsedDocument | None:
    """ZIP 容器本身不是 xlsx/docx 时，逐个尝试包内的文档文件（岗位表常在包内）。"""
    try:
        archive = zipfile.ZipFile(io.BytesIO(body))
    except Exception:
        return None
    best: ParsedDocument | None = None
    texts: list[str] = []
    for name in archive.namelist():
        lower = name.lower()
        if not lower.endswith((".xlsx", ".xlsm", ".xls", ".docx", ".doc", ".txt", ".csv")):
            continue
        if any(part.startswith("__MACOSX") or part.startswith(".") for part in name.split("/")):
            continue
        try:
            entry_body = archive.read(name)
        except Exception:
            continue
        parsed = parse_document(entry_body, "", name)
        if parsed.parser_status not in {"parsed", "empty"} or not (parsed.rows or parsed.text):
            continue
        if parsed.rows and (best is None or not best.rows):
            best = ParsedDocument(parsed.text, rows=parsed.rows, sheets=parsed.sheets, warnings=[f"解析自 ZIP 包内文件: {name}"])
        elif parsed.text:
            texts.append(f"[{name}]\n{parsed.text}")
    if best is not None:
        return best
    if texts:
        return ParsedDocument("\n\n".join(texts), warnings=["解析自 ZIP 包内文本文件"])
    return None


def _parse_ole2(body: bytes) -> ParsedDocument:
    spreadsheet = _parse_xls(body)
    if spreadsheet.parser_status in {"parsed", "empty"} and (spreadsheet.rows or spreadsheet.text):
        return ParsedDocument(spreadsheet.text, rows=spreadsheet.rows, sheets=spreadsheet.sheets, parser_status=spreadsheet.parser_status, warnings=["按魔数识别为 OLE2 复合文档（xls）"] + spreadsheet.warnings)
    word = _parse_doc(body, note="按魔数识别为 OLE2 复合文档（doc）")
    if word.parser_status in {"parsed", "empty"} and word.text:
        return word
    return ParsedDocument("", parser_status="unsupported_format", warnings=["OLE2 复合文档无法解析为 xls/doc，请下载附件查看"])


def _parse_doc(body: bytes, note: str = "") -> ParsedDocument:
    """提取旧版 Word（.doc）二进制文本：通过 FIB 定位 Clx/piece table。"""
    try:
        import olefile  # type: ignore
    except ImportError:
        return ParsedDocument("", parser_status="needs_dependency", warnings=["DOC parsing requires olefile", note])
    try:
        archive = olefile.OleFileIO(io.BytesIO(body))
        word_stream = archive.openstream("WordDocument").read()
        flags = int.from_bytes(word_stream[0x000A:0x000C], "little")
        table_name = "1Table" if flags & 0x0200 else "0Table"
        table_stream = archive.openstream(table_name).read()
        fc_clx = int.from_bytes(word_stream[0x01A2:0x01A6], "little")
        lcb_clx = int.from_bytes(word_stream[0x01A6:0x01AA], "little")
        text = _extract_doc_text(word_stream, table_stream[fc_clx:fc_clx + lcb_clx])
    except Exception as exc:
        return ParsedDocument("", parser_status="error", warnings=[f"DOC 解析失败：{exc}", note])
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", text)
    text = re.sub(r"\r+", "\n", text)
    text = "\n".join(line.strip() for line in text.splitlines() if line.strip())
    if not text:
        return ParsedDocument("", parser_status="empty", warnings=["DOC 无可提取文本", note])
    return ParsedDocument(text, warnings=[note] if note else [])


def _extract_doc_text(word_stream: bytes, clx: bytes) -> str:
    """按 piece table（PlcPcd）拼接 WordDocument 流中的文本块。"""
    index = 0
    plc_pcd: bytes | None = None
    while index < len(clx):
        block_type = clx[index]
        if block_type == 0x01:
            size = int.from_bytes(clx[index + 1:index + 3], "little")
            index += 3 + size
        elif block_type == 0x02:
            size = int.from_bytes(clx[index + 1:index + 5], "little")
            plc_pcd = clx[index + 5:index + 5 + size]
            break
        else:
            raise ValueError(f"未知 Clx 块类型: {block_type}")
    if plc_pcd is None or len(plc_pcd) < 12:
        raise ValueError("缺少 piece table")
    piece_count = (len(plc_pcd) - 4) // 12
    offsets = [int.from_bytes(plc_pcd[i * 4:i * 4 + 4], "little") for i in range(piece_count + 1)]
    parts: list[str] = []
    for piece_index in range(piece_count):
        pcd = plc_pcd[4 * (piece_count + 1) + piece_index * 8:4 * (piece_count + 1) + (piece_index + 1) * 8]
        fc_raw = int.from_bytes(pcd[2:6], "little")
        length = offsets[piece_index + 1] - offsets[piece_index]
        if length <= 0:
            continue
        if fc_raw & 0x40000000:
            start = (fc_raw & 0x3FFFFFFF) // 2
            parts.append(word_stream[start:start + length].decode("cp1252", errors="replace"))
        else:
            start = fc_raw & 0x3FFFFFFF
            parts.append(word_stream[start:start + length * 2].decode("utf-16-le", errors="replace"))
    return "".join(parts)


def _looks_binary(body: bytes) -> bool:
    sample = body[:4096]
    return bool(sample) and b"\x00" in sample


def _parse_xls(body: bytes, note: str = "") -> ParsedDocument:
    try:
        import xlrd  # type: ignore
    except ImportError:
        return ParsedDocument("", parser_status="needs_dependency", warnings=["XLS parsing requires xlrd", note])
    try:
        book = xlrd.open_workbook(file_contents=body)
    except Exception as exc:
        return ParsedDocument("", parser_status="error", warnings=[f"XLS 解析失败：{exc}", note])
    sheets: list[tuple[str, list[list[str]]]] = []
    rows: list[list[str]] = []
    for sheet in book.sheets():
        sheet_rows: list[list[str]] = []
        for row_index in range(sheet.nrows):
            row = []
            for value in sheet.row_values(row_index):
                if isinstance(value, float) and value.is_integer():
                    row.append(str(int(value)))
                else:
                    row.append(str(value).strip())
            if any(row):
                sheet_rows.append(row)
        if sheet_rows:
            sheets.append((sheet.name, sheet_rows))
            rows.extend(sheet_rows)
    text = "\n".join("\t".join(row) for row in rows)
    return ParsedDocument(text, rows=rows, sheets=sheets, warnings=[note] if not rows else [])


def _decode_text(body: bytes, charset: str | None = None) -> str:
    for encoding in (charset, "utf-8", "gb18030", "gbk"):
        if not encoding:
            continue
        try:
            return body.decode(encoding)
        except (LookupError, UnicodeDecodeError):
            continue
    return body.decode("utf-8", errors="replace")
